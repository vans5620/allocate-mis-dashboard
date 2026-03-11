"""
Run this script weekly to regenerate data.json from the latest IPS file.
Usage: python3 update_data.py
"""
import openpyxl, json, datetime
from pathlib import Path

IPS_FILE = Path(__file__).parent.parent / "03 Client Repository IPS.xlsx"
OUT_FILE = Path(__file__).parent / "data.json"

SCHEME_MAP = {
    'Ionic Allocate Portfolio Aggressive': 'Allocate Aggressive',
    'Ionic Allocate Portfolio Moderate':   'Allocate Moderate',
    'Ionic Allocate Portfolio Equity':     'Allocate Equity',
    'Ionic Liquid Approach DPMS':          'Liquid DPMS',
    'Ionic Allocate Select Portfolio':     'Allocate Select',
    'Ionic Large Value Portfolio':         'Large Value Portfolio',
    'Ionic Co-pilot Strategy':             'Co-pilot',
    'Ionic Copilot Strategy':              'Co-pilot',
    'Ionic Summit Portfolio Aggressive':   'Summit Aggressive',
}
DPMS  = ['Allocate Aggressive','Allocate Moderate','Allocate Equity','Liquid DPMS','Summit Aggressive']
NDPMS = ['Allocate Select','Large Value Portfolio','Co-pilot']

RM_ALIASES = {
    'Karan Chandok':   ['karan chandok','karan chandhok'],
    'Prateek Chhabra': ['prateek chhabra','prateek chabra'],
    'Mani Sawhney':    ['mani sawhney','mani sawheny'],
    'Ishan Mishra':    ['ishan mishra','ishan mishra '],
}
ALIAS = {a.lower().strip(): canon for canon, lst in RM_ALIASES.items() for a in lst}

def norm_rm(raw):
    if not raw: return 'Unknown'
    k = str(raw).strip().lower()
    return ALIAS.get(k, str(raw).strip())

def map_trans(t):
    if t in ('New Activation','Initial Inflow'): return 'New Activation'
    if t in ('Top-Up','Top - Up'):               return 'Top-up'
    return None

def parse_amt(v):
    if isinstance(v,(int,float)): return float(v)
    if isinstance(v,str):
        try: return float(v.replace(',',''))
        except: return None
    return None

wb    = openpyxl.load_workbook(IPS_FILE, read_only=True, data_only=True)
ws    = wb['Client Summary']
ws_rm = wb['RM Mapping']

rm_lead = {}
for r in ws_rm.iter_rows(values_only=True):
    if r[0] and r[1]: rm_lead[norm_rm(r[0])] = str(r[1]).strip()

records = []
for r in ws.iter_rows(values_only=True):
    if not r[0]: continue
    if r[3] not in SCHEME_MAP: continue
    trans = map_trans(r[5])
    if not trans: continue
    amt = parse_amt(r[6])
    if amt is None or amt <= 0: continue
    date = r[7]
    if not hasattr(date, 'year'): continue
    div = r[16] if r[16] in ('HNI','UHNI') else 'Other'
    rm  = norm_rm(r[13])
    ml  = rm_lead.get(rm, str(r[15]).strip() if r[15] else 'Unknown')
    records.append({
        'month': date.strftime('%Y-%m'), 'scheme': SCHEME_MAP[r[3]],
        'trans_type': trans, 'amount': amt, 'rm': rm,
        'market_lead': ml, 'division': div,
        'client_code': str(r[1]) if r[1] else None
    })

out = {
    'records': records, 'dpms': DPMS, 'ndpms': NDPMS,
    'all_schemes': DPMS + NDPMS,
    'last_updated': datetime.date.today().isoformat()
}
with open(OUT_FILE, 'w') as f: json.dump(out, f)
total = sum(r['amount'] for r in records)
print(f"✅  data.json updated — {len(records)} records, ₹{total/1e7:.2f} Cr, date: {out['last_updated']}")
